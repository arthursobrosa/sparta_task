from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import os
from tqdm import tqdm
import shutil
from typing import Literal
from .utils import get_suffix, get_json_data
from .sheet_info import get_tab, get_value_at_coordinate
from .distributor_info import get_distributor_info, get_column_info
from .cover_info import get_process_year, get_contract_type
from .specific_info import get_specific_info, get_real_OM_headers


def move_misplaced_files():
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    distributors_path = os.path.join(base_path, "Distribuidoras")

    distributors = [
        name for name in os.listdir(distributors_path)
        if os.path.isdir(os.path.join(distributors_path, name))
    ]

    for distributor in tqdm(distributors, desc="Processando distribuidoras..."):
        distributor_path = os.path.join(distributors_path, distributor)

        for type in ["Reajuste", "Revisão"]:
            type_path = os.path.join(distributor_path, type)

            file_names = [
                name for name in os.listdir(type_path)
                if (name.endswith(".xlsx") or name.endswith(".xlsm")) 
                and not name.startswith("~$")
            ]

            for file_name in tqdm(file_names, desc=f"{distributor} - {type}", leave=False):
                file_path = os.path.join(type_path, file_name)

                try:
                    workbook = load_workbook(file_path, keep_links=False, read_only=True, data_only=True)
                except FileNotFoundError:
                    print(f"Planilha {file_path} não encontrada")
                    continue

                cover = get_tab("CAPA", workbook)

                if cover:
                    file_concession_id = get_value_at_coordinate('C23', cover)

                    if not file_concession_id:
                        file_concession_id = get_value_at_coordinate('M2', cover)

                        if not file_concession_id:
                            print(f"Planilha {file_path} com id_concessao não encontrada")
                            continue

                    aimed_concession_id = get_column_info(
                        unknown_column_name="ID CONCESSÃO",
                        known_column_name="SIGLA",
                        known_value=distributor
                    )

                    if file_concession_id != aimed_concession_id:
                        print(f"Planilha {file_path} deveria ter id_concessao {aimed_concession_id}, mas tem id_concessao {file_concession_id}")

                        right_distributor = get_column_info(
                            unknown_column_name="SIGLA",
                            known_column_name="ID CONCESSÃO",
                            known_value=file_concession_id
                        )

                        right_distributor_path = os.path.join(distributors_path, right_distributor, type, file_name)

                        if os.path.exists(right_distributor_path):
                            print(f"O caminho {right_distributor_path} já existe. Removendo {file_path}")

                            os.remove(file_path)
                            continue

                        print(f"Movendo para {right_distributor_path}")

                        shutil.move(file_path, right_distributor_path)


def _get_fixed_tab(workbook: Workbook, distributor: str, type: Literal["Reajuste", "Revisão"]) -> Worksheet:
    distributor_info = get_distributor_info(distributor)

    process_year = get_process_year(workbook)
    distributor_info["Ano"] = process_year

    contract_type = get_contract_type(workbook)
    distributor_info["Contrato"] = contract_type

    distributor_info["Tipo de Processo"] = type

    distributor_header = list(distributor_info.keys())
    distributor_values = list(distributor_info.values())

    new_workbook = Workbook()
    new_tab = new_workbook.active

    new_tab.insert_rows(0)
    new_tab.insert_rows(0)

    new_tab.append(distributor_header)
    new_tab.append(distributor_values)

    return new_tab


def _get_changing_tab(
    distributor: str, 
    workbook: Workbook, 
    type: Literal["Reajuste", "Revisão"]
) -> Worksheet:
    new_workbook = Workbook()
    new_tab = new_workbook.active

    json_path = os.path.join(os.path.dirname(__file__), "details.json")
    json_data = get_json_data(json_path)

    headers = json_data['headers']

    useful_header = None

    for header in headers:
        all_items = []
        repeats = header['repeats']
        items = header['items']
        
        if repeats:
            all_repetitions = header['repetitions']

            for index in range(len(items)):
                repetitions = all_repetitions[index]
                item = items[index]

                for _ in range(repetitions):
                    all_items.append(item)
        else:
            useful_header = header

            for item in items:
                all_items.append(item)

        new_tab.append(all_items)

    changing_values = _get_changing_values(
        distributor=distributor,
        workbook=workbook,
        header=useful_header,
        type=type
    )

    new_tab.append(changing_values)

    return new_tab


def _get_changing_values(
    distributor: str,
    workbook: Workbook, 
    header: any, 
    type: Literal["Reajuste", "Revisão"]
) -> list[any]:
    tabs = header['tabs']
    coordinates = header['coordinates']
    all_values = []

    for tab_index in range(len(tabs)):
        if 0 <= tab_index < 81:
            if type == "Reajuste":
                all_values.append("-")
                continue

        tab_name = tabs[tab_index]

        if tab_name == "":
            specific_tabs_indexes = [81, 82, 83, 84, 85, 86, 96]

            if tab_index in specific_tabs_indexes:
                distributor_contract_type = get_column_info(
                    unknown_column_name="CONTRATO",
                    known_column_name="SIGLA",
                    known_value=distributor
                )

                if not distributor_contract_type or not isinstance(distributor_contract_type, str):
                    distributor_contract_type = ""

                process_year = get_process_year(workbook)

                value = get_specific_info(
                    workbook=workbook,
                    contract_type=distributor_contract_type,
                    type=type,
                    process_year=process_year,
                    tab_index=tab_index
                )

                if not value:
                    value = "-"

                all_values.append(value)
                continue
            else:
                all_values.append("NA")
                continue

        try:
            tab = workbook[tab_name]
        except Exception as error:
            print(f"Could not find {tab_name} on {workbook}: {str(error)}")
            continue

        coordinate = coordinates[tab_index]

        if isinstance(coordinate, str) and coordinate.startswith("ADD"):
            parts = coordinate.split("-")
            codes = parts[1:]
            value = 0

            for code in codes:
                code_value = get_value_at_coordinate(code, tab)

                if isinstance(code_value, (float, int)):
                    value += code_value
        else:
            value = get_value_at_coordinate(coordinate, tab)

        all_values.append(value)

    return all_values


def _get_other_changing_tab(workbook: Workbook, type: Literal["Reajuste", "Revisão"]):
    json_path = os.path.join(os.path.dirname(__file__), "real_OM_info.json")
    json_data = get_json_data(json_path)

    real_OM_headers = get_real_OM_headers(
        workbook=workbook,
        json_data=json_data,
        type=type
    )

    new_workbook = Workbook()
    new_tab = new_workbook.active

    new_tab.append(real_OM_headers[0])
    new_tab.append(real_OM_headers[1])
    new_tab.append(real_OM_headers[2])
    new_tab.append(real_OM_headers[3])

    return new_tab


def _filtered_workbook(
        workbook: Workbook, 
        distributor: str, 
        type: Literal["Reajuste", "Revisão"]
) -> Workbook:
    fixed_tab = _get_fixed_tab(
        workbook=workbook,
        distributor=distributor,
        type=type
    )

    changing_tab = _get_changing_tab(
        distributor=distributor,
        workbook=workbook,
        type=type
    )

    start_column = 9

    for row in changing_tab.iter_rows(
        min_row=1, 
        max_row=changing_tab.max_row, 
        min_col=1, 
        max_col=changing_tab.max_column
    ):
        for cell in row:
            new_column = cell.column + start_column
            fixed_tab.cell(row=cell.row, column=new_column).value = cell.value

    other_changing_tab = _get_other_changing_tab(
        workbook=workbook,
        type=type
    )

    start_column += 103

    for row in other_changing_tab.iter_rows(
        min_row=1,
        max_row=other_changing_tab.max_row,
        min_col=1,
        max_col=other_changing_tab.max_column
    ):
        for cell in row:
            new_column = cell.column + start_column
            fixed_tab.cell(row=cell.row, column=new_column).value = cell.value

    new_workbook = Workbook()
    new_tab = new_workbook.active
    new_tab.title = "BANCO DE DADOS"

    for row in fixed_tab.iter_rows(
        min_row=1,
        max_row=fixed_tab.max_row,
        min_col=1,
        max_col=fixed_tab.max_column
    ):
        for cell in row:
            new_tab.cell(row=cell.row, column=cell.column).value = cell.value

    return new_workbook


def _add_header_rows(header_rows: list[any], to_sheet: Worksheet):
    worksheet = to_sheet

    for row_offset, row in enumerate(header_rows, start=1):
        for column_offset, cell in enumerate(row, start=1):
            target_row = row_offset
            target_column = column_offset

            value = getattr(cell, "value", None)
            worksheet.cell(row=target_row, column=target_column).value = value


def _sorted_workbooks(workbooks: list[Workbook]) -> list[Workbook]:
    return sorted(
        workbooks,
        key=lambda workbook: workbook.active['G4'].value
    )


def _mix_db_files(file_paths: list[str], output_name: str, sort_workbooks: bool = True):
    if not file_paths:
        print(f"Lista de caminhos de arquivos vazia (iria para {output_name})")
        return
    
    file_workbooks = [
        load_workbook(file_path, keep_links=False, read_only=True, data_only=True)
        for file_path in file_paths
    ]

    if sort_workbooks:
        file_workbooks = _sorted_workbooks(file_workbooks)

    max_row_per_sheet = 1048576

    output_workbook = Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = "BANCO DE DADOS"

    current_sheet = output_worksheet
    current_row_count = 3
    sheet_index = 0

    header_rows = [] 

    for file_index, file_workbook in enumerate(file_workbooks):
        file_worksheet = file_workbook.active

        if file_index == 0:
            header_rows = list(file_worksheet.iter_rows(min_row=1, max_row=3, values_only=False))
            _add_header_rows(header_rows, to_sheet=current_sheet)

        min_row = 4
        max_row = file_worksheet.max_row

        for row in file_worksheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
            if current_row_count >= max_row_per_sheet:
                sheet_index += 1
                new_sheet_title = f"BANCO DE DADOS - Ext {sheet_index}"
                current_sheet = output_workbook.create_sheet(title=new_sheet_title)
                _add_header_rows(header_rows, to_sheet=current_sheet)
                current_row_count = 3

            current_sheet.append(row)
            current_row_count += 1

    output_workbook.save(output_name)


def process_distributors():
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    distributors_path = os.path.join(base_path, "Distribuidoras")

    distributors = [
        name for name in os.listdir(distributors_path)
        if os.path.isdir(os.path.join(distributors_path, name))
    ]

    distributors.sort()

    for distributor in tqdm(distributors, desc="Processando distribuidoras..."):
        distributor_path = os.path.join(distributors_path, distributor)

        temp_file_paths = []

        for type in ["Reajuste", "Revisão"]:
            type_path = os.path.join(distributor_path, type)

            file_names = [
                name for name in os.listdir(type_path)
                if (name.endswith(".xlsx") or name.endswith(".xlsm")) 
                and not name.startswith("~$")
            ]

            for file_name in tqdm(file_names, desc=f"{distributor} - {type}", leave=False):
                file_path = os.path.join(type_path, file_name)
                file_workbook = load_workbook(file_path, keep_links=False, read_only=True, data_only=True)

                try:
                    new_workbook = _filtered_workbook(
                        workbook=file_workbook,
                        distributor=distributor,
                        type=type
                    )

                    file_suffix = get_suffix(file_name)
                    temp_path = file_path.replace(file_suffix, f"_temp{file_suffix}")
                    new_workbook.save(temp_path)    
                    temp_file_paths.append(temp_path)   
                except Exception as error:
                    print(f"\nFalha ao filtrar planilha em {file_path}: {str(error)}")    

        if temp_file_paths:
            output_folder_path = os.path.join(distributor_path, "Banco de Dados")
            os.makedirs(output_folder_path, exist_ok=True)   

            output_path = os.path.join(output_folder_path, f"{distributor}_BANCO.xlsx")

            _mix_db_files(
                file_paths=temp_file_paths,
                output_name=output_path
            )

            print(f"\nBanco de dados consolidado em {output_path}")

            for temp_file in temp_file_paths:
                os.remove(temp_file)


def process_data_bases():
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    distributors_path = os.path.join(base_path, "Distribuidoras")

    distributors = [
        name for name in os.listdir(distributors_path)
        if os.path.isdir(os.path.join(distributors_path, name))
    ]

    distributors.sort()

    file_paths = []

    for distributor in tqdm(distributors, desc="Processando distribuidoras..."):
        distributor_path = os.path.join(distributors_path, distributor)

        db_path = os.path.join(distributor_path, "Banco de Dados")

        file_names = [
            name for name in os.listdir(db_path)
            if (name.endswith(".xlsx") or name.endswith(".xlsm")) 
            and not name.startswith("~$")
        ]

        for file_name in tqdm(file_names, desc=f"{distributor} - {type}", leave=False):
            file_path = os.path.join(db_path, file_name)
            file_paths.append(file_path)

    if file_paths:
        output_path = os.path.join(base_path, "BANCO.xlsx")

        _mix_db_files(
            file_paths=file_paths,
            output_name=output_path,
            sort_workbooks=False
        )

        print(f"\nBanco de dados consolidado em {output_path}")


def remove_dbs():
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    distributors_path = os.path.join(base_path, "Distribuidoras")

    distributors = [
        name for name in os.listdir(distributors_path)
        if os.path.isdir(os.path.join(distributors_path, name))
    ]

    distributors.sort()

    for distributor in distributors:
        distributor_path = os.path.join(distributors_path, distributor)
        db_path = os.path.join(distributor_path, "Banco de Dados")

        file_names = [
            name for name in os.listdir(db_path)
            if (name.endswith(".xlsx") or name.endswith(".xlsm")) 
            and not name.startswith("~$")
        ]

        for file_name in file_names:
            file_path = os.path.join(db_path, file_name)
            os.remove(file_path)

        os.rmdir(db_path)