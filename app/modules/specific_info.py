from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from typing import Literal
from .cover_info import get_value_at_coordinate
from .utils import normalize
import os



def _get_detailed_info(
    workbook: Workbook,
    default_tab_name: str,
    default_coordinate: str,
    alternate_tab_name: str,
    min_row: int,
    offset_to_max_row: int,
    first_column_index: int,
    second_column_index: int,
    target_value: str,
    second_target_value: str,
    column_offset: int,
    first_coordinate_index: int = 0,
    second_coordinate_index: int = 0
):
    try:
        tab = workbook[default_tab_name]
        value = get_value_at_coordinate(default_coordinate, tab)
        return value
    except Exception:
        pass

    try:
        tab = workbook[alternate_tab_name]
    except Exception:
        return None
    
    first_coordinates = []
    first_coordinate = ""
    second_coordinates = []
    second_coordinate = ""

    for row in tab.iter_rows(min_row=min_row, min_col=first_column_index, max_col=first_column_index):
        for cell in row:
            value = cell.value

            if isinstance(value, str) and normalize(value) == target_value:
                first_coordinates.append(cell.coordinate)

    first_coordinates = sorted(
        first_coordinates,
        key=lambda coordinate: coordinate_from_string(coordinate)[1]
    )

    if len(first_coordinates) > 0:
        first_coordinate = first_coordinates[first_coordinate_index]
    else:
        return
    
    _, first_row_index = coordinate_from_string(first_coordinate)
    max_row = first_row_index + offset_to_max_row

    for row in tab.iter_rows(min_row=first_row_index, max_row=max_row, min_col=second_column_index, max_col=second_column_index):
        for cell in row:
            value = cell.value

            if isinstance(value, str) and normalize(value) == second_target_value:
                second_coordinates.append(cell.coordinate)
                break

    second_coordinates = sorted(
        second_coordinates,
        key=lambda coordinate: coordinate_from_string(coordinate)[1]
    )

    if len(second_coordinates) > 0:
        second_coordinate = second_coordinates[second_coordinate_index]
    else:
        return
    
    second_column_letter, second_row_index = coordinate_from_string(second_coordinate)
    second_column_index = column_index_from_string(second_column_letter)

    second_column_index += column_offset
    second_column_letter = get_column_letter(second_column_index)
    
    value_coordinate = f"{second_column_letter}{second_row_index}"

    return get_value_at_coordinate(value_coordinate, tab)


def _get_OR_info(
    workbook: Workbook,
    contract_type: str,
    type: Literal["Reajuste", "Revisão"]
):
    if type == "Revisão":
        return _get_detailed_info(
            workbook=workbook,
            default_tab_name="UDEROR",
            default_coordinate='C2',
            alternate_tab_name="VPB e Fator X",
            min_row=160,
            offset_to_max_row=30,
            first_column_index=2,
            second_column_index=2,
            target_value="FATOR IGP- M",
            second_target_value="TOTAL",
            column_offset=2
        )
    else:
        if contract_type == "Novo":
            return _get_detailed_info(
                workbook=workbook,
                default_tab_name="UDEROR",
                default_coordinate='C2',
                alternate_tab_name="VPB1",
                min_row=36,
                offset_to_max_row=0,
                first_column_index=2,
                second_column_index=2,
                target_value="OUTRAS RECEITAS (OR)",
                second_target_value="OUTRAS RECEITAS (OR)",
                column_offset=1
            )
        else:
            return None


def _get_UD_or_ER_info(
    ud_or_er: Literal["UD", "ER"],
    workbook: Workbook,
    contract_type: str,
    type: Literal["Reajuste", "Revisão"],
    process_year: any
):
    if not isinstance(process_year, int):
        return None

    if type == "Revisão":
        if contract_type == "Novo":
            is_old = process_year < 2021
        elif contract_type == "Antigo":
            is_old = process_year < 2019
        else:
            return None
        
        if is_old:
            return None
        
        target_value = "ULTRAPASSAGEM DE DEMANDA" if ud_or_er == 'UD' else "EXCEDENTE DE REATIVOS"
        
        return _get_detailed_info(
            workbook=workbook,
            default_tab_name="",
            default_coordinate="",
            alternate_tab_name="VPB e Fator X",
            min_row=160,
            offset_to_max_row=0,
            first_column_index=2,
            second_column_index=2,
            target_value=target_value,
            second_target_value=target_value,
            column_offset=2
        )
    elif type == "Reajuste":
        if contract_type != "Novo" or process_year < 2017:
            return None
        
        default_coordinate = 'C6' if ud_or_er == "UD" else 'C7'

        target_value = "ULTRAPASSAGEM DE DEMANDA (UD)" if ud_or_er == "UD" else "EXCEDENETE DE REATIVOS (ER)"      
        
        return _get_detailed_info(
            workbook=workbook,
            default_tab_name="UDEROR",
            default_coordinate=default_coordinate,
            alternate_tab_name="VPB1",
            min_row=36,
            offset_to_max_row=0,
            first_column_index=2,
            second_column_index=2,
            target_value=target_value,
            second_target_value=target_value,
            column_offset=1
        )
    
    return None


def _get_dr1_factor_info(
    workbook: Workbook,
    contract_type: str,
    type: Literal["Reajuste", "Revisão"],
    process_year: any
):
    if not isinstance(process_year, int):
        return None

    if contract_type == "Novo" and type == "Reajuste" and process_year >= 2017:
        try:
            tab = workbook["BD"]
            return get_value_at_coordinate('M63', tab)
        except Exception:
            pass

        try:
            tab = workbook["VPB1"]
            return get_value_at_coordinate('C38', tab)
        except Exception:
            return None
    else:
        return None
    

def _get_enc_RI_info(
    workbook: Workbook,
    type: Literal["Reajuste", "Revisão"],
):
    if type == "Revisão":
        return _get_detailed_info(
            workbook=workbook,
            default_tab_name="",
            default_coordinate="",
            alternate_tab_name="VPB e Fator X",
            min_row=180,
            offset_to_max_row=0,
            first_column_index=2,
            second_column_index=2,
            target_value="RECEITAS IRRECUPERAVEIS DE ENCARGOS SETORIAIS",
            second_target_value="RECEITAS IRRECUPERAVEIS DE ENCARGOS SETORIAIS",
            column_offset=1
        )
    else:
        return None
    

def _get_dem_RI_info(
    workbook: Workbook,
    type: Literal["Reajuste", "Revisão"],
    process_year: any
):
    if type == "Revisão":
        return _get_detailed_info(
            workbook=workbook,
            default_tab_name="",
            default_coordinate="",
            alternate_tab_name="VPB e Fator X",
            min_row=190,
            offset_to_max_row=0,
            first_column_index=2,
            second_column_index=2,
            target_value="DEMAIS RECEITAS IRRECUPERAVEIS (VSE)",
            second_target_value="DEMAIS RECEITAS IRRECUPERAVEIS (VSE)",
            column_offset=1
        )
    
    if not isinstance(process_year, int):
        return None

    if type == "Reajuste" and process_year >= 2017:
        return _get_detailed_info(
            workbook=workbook,
            default_tab_name="Mercado",
            default_coordinate="H38",
            alternate_tab_name="Resultado",
            min_row=30,
            offset_to_max_row=0,
            first_column_index=2,
            second_column_index=2,
            target_value="RECEITA IRRECUPERAVEL",
            second_target_value="RECEITA IRRECUPERAVEL",
            column_offset=2
        )
    else:
        return None


def _get_pb_info(
    workbook: Workbook,
    type: Literal["Reajuste", "Revisão"]
):
    target_value = "PARCELA B (MENOS OUTRAS RECEITAS)" if type == "Revisão" else "PARCELA B"
    first_coordinate_index = 0 if type == "Revisão" else 1

    return _get_detailed_info(
        workbook=workbook,
        default_tab_name="",
        default_coordinate="",
        alternate_tab_name="Resultado",
        min_row=38,
        offset_to_max_row=0,
        first_column_index=2,
        second_column_index=2,
        target_value=target_value,
        second_target_value=target_value,
        column_offset=2,
        first_coordinate_index=first_coordinate_index
    )


def get_specific_info(
    workbook: Workbook,
    contract_type: str,
    type: Literal["Reajuste", "Revisão"],
    process_year: any,
    tab_index: int
):
    if tab_index == 81:
        return _get_OR_info(
            workbook=workbook,
            contract_type=contract_type,
            type=type
        )
    
    if tab_index == 82:
         return _get_UD_or_ER_info(
            ud_or_er='UD',
            workbook=workbook,
            contract_type=contract_type,
            type=type,
            process_year=process_year
        )
    
    if tab_index == 83:
        return _get_UD_or_ER_info(
            ud_or_er='ER',
            workbook=workbook,
            contract_type=contract_type,
            type=type,
            process_year=process_year
        )
    
    if tab_index == 84:
        return _get_dr1_factor_info(
            workbook=workbook,
            contract_type=contract_type,
            type=type,
            process_year=process_year
        )
    
    if tab_index == 85:
        return _get_enc_RI_info(
            workbook=workbook,
            type=type
        )
    
    if tab_index == 86:
        return _get_dem_RI_info(
            workbook=workbook,
            type=type,
            process_year=process_year
        )
    
    if tab_index == 96:
        return _get_pb_info(
            workbook=workbook,
            type=type
        )


def _get_t1_and_t2_dict(workbook: Workbook):
    try:
        tab = workbook["Entrada"]
    except Exception:
        return None
    
    t1_and_t2 = {}
    coordinates = ['L7', 'M7']

    for coordinate in coordinates:
        value = get_value_at_coordinate(coordinate, tab)

        if value and isinstance(value, int):
            t1_and_t2[coordinate] = value

    t1_and_t2 = dict(
        sorted(
            t1_and_t2.items(),
            key=lambda item: item[1]
        )
    )

    return t1_and_t2


def _get_t_column_index(t: int, t_coordinates: list[str]):
    t1_coordinate = t_coordinates[0]
    t2_coordinate = t_coordinates[1]

    t_coordinate = t1_coordinate if t == 0 else t2_coordinate
    t_column_letter, _ = coordinate_from_string(t_coordinate)
    return column_index_from_string(t_column_letter)


def get_real_OM_headers(
    workbook: Workbook,
    json_data: any,
    type: Literal["Reajuste", "Revisão"]
):
    base_name = json_data['base_name']
    items = json_data['items']
    sub_items = json_data['sub_items']

    t1_and_t2_dict = _get_t1_and_t2_dict(workbook)

    if t1_and_t2_dict and len(t1_and_t2_dict) == 2:
        coordinates = list(t1_and_t2_dict.keys())
        years = list(t1_and_t2_dict.values())
    else:
        coordinates = None
        years = None

    t1_year = years[0] if years else 0
    t2_year = years[1] if years else 0

    right_coordinates = coordinates if coordinates else []

    is_empty = (t1_year == 0 and t2_year == 0 and right_coordinates == []) or type == "Reajuste"

    first_header = []
    second_header = []
    third_header = []
    values = []

    for t in range(2):
        t_text = "t-1" if t == 0 else "t-2"

        for i in range(len(items)):
            main_item = items[i]
            minor_items = sub_items[i]

            second_header_item = f"{base_name} {main_item} {t_text}"

            for j in range(len(minor_items)):
                first_header.append("Revisão")
                second_header.append(second_header_item)

                minor_item = f"{minor_items[j]} {t_text}" if (i == 0 and j == 0) else minor_items[j]
                third_header.append(minor_item)

                if is_empty:
                    values.append("-")
                    continue

                t_year = t1_year if t == 0 else t2_year

                target_value = normalize(main_item) if i == len(items) - 1 else f"CONTA {normalize(main_item)}"
                t_column_index = _get_t_column_index(t, coordinates)

                value = (t_year if (i == 0 and j == 0) else
                    _get_detailed_info(
                        workbook=workbook,
                        default_tab_name="",
                        default_coordinate="",
                        alternate_tab_name="Entrada",
                        min_row=7,
                        offset_to_max_row=32,
                        first_column_index=10,
                        second_column_index=11,
                        target_value=target_value,
                        second_target_value=normalize(minor_item),
                        column_offset=t_column_index - 11
                    )
                )

                values.append(value)

    return [
        first_header,
        second_header,
        third_header,
        values
    ]