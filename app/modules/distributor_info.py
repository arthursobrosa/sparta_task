from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os


def load_distributors_sheet() -> Worksheet:
    file_path = os.path.join(os.path.dirname(__file__), "../../distribuidoras.xlsx")
    file_path = os.path.abspath(file_path)

    workbook = load_workbook(file_path, keep_links=False, read_only=True, data_only=True)
    return workbook.active


def get_distributor_info(distributor: str):
    name = get_column_info(
        unknown_column_name="NOME",
        known_column_name="SIGLA",
        known_value=distributor
    )

    agent = get_column_info(
        unknown_column_name="AGENTE",
        known_column_name="SIGLA",
        known_value=distributor
    )

    concession_id = get_column_info(
        unknown_column_name="ID CONCESSÃO",
        known_column_name="SIGLA",
        known_value=distributor
    )

    code = get_column_info(
        unknown_column_name="CÓDIGO",
        known_column_name="SIGLA",
        known_value=distributor
    )

    agent_id = get_column_info(
        unknown_column_name="ID AGENTE",
        known_column_name="SIGLA",
        known_value=distributor
    )

    return {
        'SIGLA': distributor,
        'NOME': name,
        'AGENTE': agent,
        'ID CONCESSÃO': concession_id,
        'CÓDIGO': code,
        'ID AGENTE': agent_id
    }



def get_column_info(unknown_column_name: str, known_column_name: str, known_value: str):
    worksheet = load_distributors_sheet()

    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    known_value_index = header.index(known_column_name)
    unknown_value_index = header.index(unknown_column_name)

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        known_cell = row[known_value_index]
        unknown_cell = row[unknown_value_index]

        if known_cell == known_value:
            return unknown_cell
        
    return None